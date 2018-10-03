import requests as req
import json
import pprint as pretty
import openpyxl as opxl
from datetime import datetime as dt
import pandas as pd

key = "7af48c8988b13c4580a58cb9ba0158c0"
url = "https://api.monday.com:443"

board_id = "120457797"

all_users = "/v1/users.json"
updates = "/v1/updates.json"
order_setup = "/v1/boards/120457797/pulses.json?per_page=25"

def get(query):
    """Use requests.get to return Monday.com api call"""
    call = url + query + "&api_key=" + key
    result = req.get(call)
    print("Trying to return {}".format(call))
    return result

response = get(order_setup)

data = json.loads(response.content)

#print(data)

test = pd.DataFrame()
test1 = pd.DataFrame()

for d in data[:3]:
    test = test.append(pd.DataFrame(d['pulse'], index=[0]))
    test1 = test1.append(pd.DataFrame(d['column_values']))

##pid = []
##name = []
##created_date = []
##customer_name = []
##
##for d in data:
##    pid.append(d['pulse']['id'])
##    name.append(d['pulse']['name'])
##    created_date.append(d['pulse']['created_at'].split("T")[0])
##
##order_status = pd.DataFrame(
##    {'OrderId': pid,
##     'CreatedDate': created_date,
##     'OrderName': name
##     })
##
### reset column order in dataframe
##order_status = order_status[['OrderId','CreatedDate','OrderName']]
##
##orders_by_day = order_status.groupby(['CreatedDate']).size()
##
# Write to Excel using Pandas
writer = pd.ExcelWriter('Output.xlsx')
test.to_excel(writer, 'OrderStatus')
test1.to_excel(writer, 'Columns')
writer.save()
##
### Write to Excel using OpenPyExcel
##wb = opxl.Workbook()
##ws = wb.active
##
##pyOut = "Test.xlsx"
##
##colHeaders = ["PID", "PO Name", "Created Date"]
##
##for col in range(0, len(colHeaders)):
##    ws.cell(column=col+1, row=1, value=colHeaders[col])
##
##try:
##    for row in range(2, len(pid)+2): # add data from lists (start at 2 because excel is unity indexed and a header exists already)
##        ws.cell(column=1, row=row, value=pid[row-2])
##        ws.cell(column=2, row=row, value=name[row-2])
##        ws.cell(column=3, row=row, value=created_date[row-2])
##
##except IndexError:
##    pass
##
##wb.save(filename = pyOut)                       
##
####pulseId = []
####pulseName = []
####
####for d in data:
####    
####    pulseId.append(d["pulse"]["id"])
####    pulseName.append(d["pulse"]["name"])
####
####print(pulseId)
####print(pulseName)
    
